# excel.py

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import os  # Windows용 CMD 일시정지

def merge_contract_price(order_file_path, price_file_path, output_dir):
    output_dir.mkdir(parents=True, exist_ok=True)
    price_df = pd.read_excel(price_file_path, header=3)
    price_df.columns = price_df.columns.astype(str).str.strip()

    price_map = dict(
        zip(
            price_df["식품 공통코드명"].astype(str).str.strip(),
            price_df["②입찰단가"]
        )
    )

    wb = load_workbook(order_file_path)
    ws = wb.active

    # NO 헤더 행 찾기
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20), 1):
        if row[0].value == "NO":
            header_row = i
            break
    if header_row is None:
        raise ValueError("❌ 헤더 행을 찾을 수 없습니다 (NO 기준)")

    # 식품명 열 찾기
    식품명_col = None
    for cell in ws[header_row]:
        if cell.value and "식품명" in str(cell.value):
            식품명_col = cell.col_idx
            break
    if 식품명_col is None:
        raise ValueError("❌ '식품명' 열을 찾을 수 없습니다.")

    # 계약단가 열 추가
    last_col = ws.max_column
    계약단가_col = last_col + 1
    ws.cell(row=header_row, column=계약단가_col, value="계약단가")

    # 각 행마다 매핑된 단가 써넣기
    for row in range(header_row + 1, ws.max_row + 1):
        식품명 = ws.cell(row=row, column=식품명_col).value
        if 식품명 and str(식품명).strip().lower() not in ["nan", ""]:
            matched_price = price_map.get(str(식품명).strip(), "")
            ws.cell(row=row, column=계약단가_col, value=matched_price)

    # 결과 저장 (파일명에 _업로드용 추가)
    output_filename = order_file_path.stem + "_업로드용.xlsx"
    output_path = output_dir / output_filename
    wb.save(output_path)
    print(f"✅ 저장 완료: {output_path}")
    return output_path

if __name__ == "__main__":
    base_dir = Path("C:/school/upload")
    output_dir = base_dir / "excel"

    # 모든 발주서 파일 처리
    order_files = list(base_dir.glob("*_발주서_*.xlsx"))
    for order_file in order_files:
        name_parts = order_file.stem.split("_")
        if len(name_parts) < 4:
            print(f"⚠️ 파일 이름 형식 오류: {order_file.name}")
            continue

        yyyymm, school = name_parts[0], name_parts[1]
        price_file = base_dir / f"{yyyymm}_{school}_계약단가.xlsx"
        if not price_file.exists():
            print(f"❌ 계약단가 파일 없음: {price_file.name}")
            continue

        try:
            merge_contract_price(order_file, price_file, output_dir)
        except Exception as e:
            print(f"❌ 처리 실패: {order_file.name} → {e}")

    # CMD 창이 바로 닫히지 않도록 일시정지
    os.system("pause")
